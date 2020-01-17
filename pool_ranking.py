import os
import pickle
import openpyxl as xl


with open('players', 'rb') as fichier:
    mon_depickler = pickle.Unpickler(fichier)
    players = mon_depickler.load()

with open('matchs', 'rb') as fichier:
    mon_depickler = pickle.Unpickler(fichier)
    matchs = mon_depickler.load()

# print(players)


results = {}
for elt in players:
    results[elt] = 0

# for i, j in results.items():
#     print(f"{i} = {j} pts")

for elt in matchs:
    print(f"{elt} : who won the match ? ")
    okay = False
    while not okay:
        winner = input(":->").title()
        if winner not in players:
            print("Unkown player")
            continue
        elif winner in players:
            points = results.get(winner)
            results[winner] = points + 3
            okay = True

wb = xl.load_workbook("poolresults.xlsx")
sheet = wb["Feuil1"]

row = 3

for i, j in results.items():
    cell_name = sheet.cell(row, 3)
    cell_marks = sheet.cell(row, 4)
    cell_name.value = i
    cell_marks.value = j
    row += 1

wb.save("poolresults.xlsx")

os.system("pause")
